
import json
import re
from pathlib import Path
from dataclasses import asdict, dataclass, field
from typing import Optional
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

@dataclass
class AnalysisResult:
    document_type: str
    project_name: Optional[str]
    project_reference: Optional[str]
    order_number: Optional[str]
    notification_number: Optional[str]
    revision: Optional[str]
    discipline: Optional[str]
    nature_of_work: str
    work_instruction_procedure: str
    inspection_mode: Optional[str]
    inspection_stage: Optional[str]
    equipment: Optional[str]
    scope_of_work: Optional[str]
    equipment_scope: Optional[str]
    inspection_factory_location: Optional[str]
    site_of_inspection: Optional[str]
    named_contacts: list[str] = field(default_factory=list)
    relevant_itp_step: Optional[str] = None
    applicable_standards: list[str] = field(default_factory=list)
    tests_or_checks: list[str] = field(default_factory=list)
    inspector_duties: list[str] = field(default_factory=list)
    referenced_documents: list[str] = field(default_factory=list)
    referenced_documents_summary: str = ""
    reporting_requirements: list[str] = field(default_factory=list)
    missing_information: list[str] = field(default_factory=list)
    confidence_level: str = "Low"
    reasoned_summary: str = ""

def clean(text):
    return re.sub(r"\s+", " ", str(text or "")).strip()

def uniq(items):
    out, seen = [], set()
    for item in items:
        c = clean(item)
        if c and c not in seen:
            out.append(c); seen.add(c)
    return out

def first(text, pattern):
    m = re.search(pattern, text, flags=re.I | re.S)
    return clean(m.group(1)) if m else None

def allm(text, pattern):
    return uniq([m.group(1) for m in re.finditer(pattern, text, flags=re.I | re.S)])

def extract_text(path: Path):
    suf = path.suffix.lower()
    if suf in {".txt", ".md", ".csv", ".log"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suf == ".docx":
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if clean(p.text))
    if suf == ".pdf":
        reader = PdfReader(str(path))
        text = "\n".join((page.extract_text() or "") for page in reader.pages[:80])
        if not clean(text):
            raise ValueError("PDF text could not be extracted. Please upload a text-based PDF or paste the extracted text.")
        return text
    if suf == ".xlsx":
        wb = load_workbook(filename=str(path), data_only=True)
        chunks = []
        for ws in wb.worksheets[:12]:
            chunks.append(f"Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                vals = [clean(v) for v in row if clean(v)]
                if vals: chunks.append(" | ".join(vals))
        text = "\n".join(chunks)
        if not clean(text): raise ValueError("Spreadsheet appears empty or unreadable.")
        return text
    raise ValueError("Unsupported file type. Supported formats: PDF, DOCX, TXT, CSV, XLSX.")

def classify(text, file_name=""):
    c = f"{text}\n{file_name}"
    if re.search(r"notification for inspection|\bNOI\b", c, re.I): return "NOI / Notification for Inspection"
    if re.search(r"notification of readiness for inspection|\bNORFI\b", c, re.I): return "NORFI / Notification of Readiness for Inspection"
    if re.search(r"\biFAT\b|system integration test", c, re.I): return "iFAT / Inspection Notification linked to ITP-QCP"
    if re.search(r"factory acceptance test plan|\bFAT\b", c, re.I): return "FAT / Factory Acceptance Test"
    if re.search(r"inspection and test plan|\bITP\b|QCP|quality control plan", c, re.I): return "Inspection and Test Plan"
    if re.search(r"call off|inspection assignment work order", c, re.I): return "Call Off / Inspection Assignment Work Order"
    if re.search(r"work instruction|procedure|instruction", c, re.I): return "Work Instruction / Procedure"
    return "Unknown"

def extract_noi_activities(text):
    activities = []
    if re.search(r"\bF47\b", text, re.I) and re.search(r"partial.?discharge", text, re.I):
        activities.append("F47 — Partial-discharge measurement")
    if re.search(r"\bF61\b", text, re.I) and re.search(r"impulse|AC voltage", text, re.I):
        activities.append("F61 — Impulse or AC voltage test on two single coils")
    return uniq(activities)

def extract_xlsx_activities(text):
    activities = []
    if re.search(r"UNIT CONTROL PANEL", text, re.I): activities.append("UNIT CONTROL PANEL — iFAT")
    if re.search(r"SYSTEM INTEGRATION TEST", text, re.I): activities.append("System Integration Test")
    for line in text.splitlines():
        c = clean(line)
        if c and re.search(r"\bBH\b.*\bW\b|\bWitness\b|\bHold\b|\bObserve\b|Inspection Date", c, re.I):
            if not re.search(r"document return code|status|revision code", c, re.I):
                activities.append(c)
    return uniq(activities)[:12]

def extract_referenced_documents(text):
    docs = []
    patterns = [
        r"(QUALITY CONTROL PLAN[^\n|]*)",
        r"(UNIT CONTROL SYSTEM SCHEMATIC[^\n|]*)",
        r"(UNIT CONTROL SYSTEM I/O LIST[^\n|]*)",
        r"(UNIT CONTROL SYSTEM INTERNAL WIRING DIAGRAM[^\n|]*)",
        r"(INTERCONNECTING WIRING DIAGRAM[^\n|]*)",
        r"(LAYOUT / CONSTRUCTION & MAIN COMPONENTS LIST[^\n|]*)",
        r"(UNIT CONTROL SYSTEM TEST PROCEDURE[^\n|]*)",
    ]
    for p in patterns: docs += allm(text, p)
    return uniq(docs)[:20]

def referenced_docs_summary(docs):
    if not docs:
        return "No additional referenced documents were identified in the uploaded instruction."
    return "Additional referenced documents appear to support later photo/evidence review, including control plans, schematics, I/O lists, wiring diagrams, layout/component lists, or test procedures. These should be used as supporting context when judging photos, measured values, terminal wiring, panel layout, and test evidence."

def extract_reporting_requirements(text):
    reqs = []
    if re.search(r"must be e-?mailed|emailed", text, re.I): reqs.append("Inspection notification must be emailed to the listed recipients.")
    if re.search(r"minimum\s+14\s+calendar\s+days", text, re.I): reqs.append("Notification must be issued at least 14 calendar days before the inspection date.")
    if re.search(r"working days in advance", text, re.I): reqs.append(first(text, r"(\d+\s+working days in advance[^\n|]{0,140})") or "Advance notice is required before inspection.")
    return uniq(reqs)

def extract_contacts(text):
    phones = []
    for p in allm(text, r"(\+?\d[\d ()-]{7,}\d)"):
        if not re.search(r"^\d{4,}$", p): phones.append(f"Phone: {p}")
    return uniq([f"Email: {m}" for m in allm(text, r"([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})")] + phones + [f"Contact person: {m}" for m in allm(text, r"Contact Person\s*(?:Section Project Manager Name)?\s*([^\n]+)")] + [f"Vendor / Subcontractor: {m}" for m in allm(text, r"Vendor\s*/\s*Sub\s*contractor Name\s*([^\n]+)")])[:16]

def analyse_text(text, file_name=""):
    doc_type = classify(text, file_name)
    standards = uniq([m.group(1) for m in re.finditer(r"(IEC(?:/IEEE)?\s*[0-9-]+(?:-[0-9]+)?)", text, flags=re.I)])[:20]
    if doc_type == "NOI / Notification for Inspection": tests = extract_noi_activities(text)
    elif "iFAT" in doc_type or file_name.lower().endswith(".xlsx"): tests = extract_xlsx_activities(text)
    else: tests = uniq([f"{m.group(1)} — {clean(m.group(2))}" for m in re.finditer(r"(?:Item\s*)?(\d+(?:\.\d+)*)\s+([^\n]{6,120})", text, flags=re.I)])[:20]
    project_ref = first(text, r"PROJECT No\.?\s*([0-9A-Z.-]+)")
    project_name = first(text, r"Project Name\s*:?\s*([^\n|]+)") or first(text, r"(MERAM PROJECT)") or first(text, r"(Trion FPU Project)")
    equipment = first(text, r"PO Description\s*([^\n]+)") or first(text, r"Equipment Description\s*&\s*Tag No\.?\s*\|?\s*([^\n|]+)") or ("UNIT CONTROL PANEL" if re.search(r"UNIT CONTROL PANEL", text, re.I) else None)
    scope = None
    if re.search(r"partial.?discharge|impulse|AC voltage", text, re.I): scope = "Witness electrical tests listed in the NOI / ITP steps."
    if re.search(r"SYSTEM INTEGRATION TEST|iFAT", text, re.I): scope = "iFAT / System Integration Test and linked QCP / ITP checks."
    location = first(text, r"(Čedlosy\s+126,\s*664\s*24\s*Drasov)") or first(text, r"(East Gate Business Park F2,\s*H-2151\s*Fót,\s*Hungary)") or first(text, r"Inspection Location\s*:?\s*([^\n|]+)") or first(text, r"Address\s*([^\n]+Drasov)")
    docs = extract_referenced_documents(text)
    nature_lookup = {
        "FAT / Factory Acceptance Test": "Factory acceptance test procedure / witness activity",
        "iFAT / Inspection Notification linked to ITP-QCP": "Work instruction / procedure linked to iFAT, QCP and ITP witness planning",
        "NOI / Notification for Inspection": "Work instruction / procedure for notification-based witness inspection",
        "NORFI / Notification of Readiness for Inspection": "Work instruction / procedure for readiness-based inspection attendance",
        "Inspection and Test Plan": "Work instruction / procedure defining inspection checkpoints and acceptance criteria",
        "Call Off / Inspection Assignment Work Order": "Work instruction / procedure defining assignment, location, parties and reporting route",
        "Work Instruction / Procedure": "Work instruction / procedure",
        "Unknown": "Instruction type not identified confidently",
    }
    result = AnalysisResult(
        document_type=doc_type, project_name=project_name,
        project_reference=project_ref or first(text, r"RFQ No\.?\s*/\s*PO No\.?\s*\|?\s*([^|\n]+)"),
        order_number=first(text, r"Order No\.?\s*:?\s*([^\n]+)") or first(text, r"PO NO:\s*([^\n]+)"),
        notification_number=first(text, r"Notification No\.?\s*([0-9A-Z.-]+)") or first(text, r"Application No\.?\s*\|?\s*([A-Z0-9-]+)"),
        revision=first(text, r"Rev\s*#?\s*:?\s*([A-Z0-9.-]+)") or first(text, r"Rev\.?\s*([A-Z0-9.-]+)"),
        discipline=("Electrical" if re.search(r"electrical|transformer|partial.?discharge|motor|voltage|impulse|unit control", text, re.I) else None),
        nature_of_work=nature_lookup.get(doc_type, "Work instruction / procedure"),
        work_instruction_procedure=nature_lookup.get(doc_type, "Work instruction / procedure"),
        inspection_mode="Witness" if re.search(r"\bWitness\b", text, re.I) else ("Hold" if re.search(r"\bHold\b", text, re.I) else None),
        inspection_stage="FAT / iFAT / witness stage" if re.search(r"\bFAT\b|\biFAT\b|Witness", text, re.I) else None,
        equipment=equipment, scope_of_work=scope, equipment_scope=equipment or scope,
        inspection_factory_location=location, site_of_inspection=None, named_contacts=extract_contacts(text),
        relevant_itp_step=", ".join(uniq(allm(text, r"\b(F\d{1,3})\b"))) or first(text, r"ITP Step No\s*([^\n]+)"),
        applicable_standards=standards, tests_or_checks=tests, referenced_documents=docs,
        referenced_documents_summary=referenced_docs_summary(docs), reporting_requirements=extract_reporting_requirements(text),
        inspector_duties=[], missing_information=[],
        confidence_level="High" if len(clean(text)) > 500 else ("Medium" if len(clean(text)) > 120 else "Low"),
        reasoned_summary="",
    )
    if doc_type == "NOI / Notification for Inspection":
        result.inspector_duties = ["Confirm attendance or waiver before the inspection date.","Witness the listed ITP activities at the stated factory / inspection location.","Verify that the correct motor / coils are under test and that the test setup matches the referenced ITP.","Record test status, deviations, remarks, and intervention status for Supplier / Contractor / Company or TPA.","Use the referenced ITP and project quality documents for exact acceptance criteria."]
        result.reasoned_summary = f"This instruction is a NOI for {result.project_name or result.project_reference}. The inspector should attend the stated location, witness the listed ITP activities, verify the test setup and identity of the equipment, record intervention status, and report any deviation or missing acceptance evidence."
    elif "iFAT" in doc_type:
        result.inspector_duties = ["Confirm the inspection date window, location, parties and required notice period.","Review the linked QCP / ITP and referenced documents before attendance.","Witness or observe the iFAT / system integration test according to the intervention matrix.","Check that evidence photos, measured values, wiring/panel references and test results are traceable to the referenced documents.","Record open points, deviations, missing documents and final inspection status."]
        result.reasoned_summary = f"This instruction is linked to iFAT / system integration testing for {result.equipment or 'the equipment'}. The inspector should prepare with the QCP/ITP, referenced drawings and test procedure, then capture evidence against each planned test or inspection activity."
    else:
        result.inspector_duties = ["Review the document and confirm the inspection scope.","Use the extracted fields to prepare the pre-inspection briefing."]
        result.reasoned_summary = "This document has been analysed and classified to support inspection planning."
    if not (result.project_name or result.project_reference): result.missing_information.append("Project name / reference not identified confidently.")
    if not result.equipment: result.missing_information.append("Equipment not identified confidently.")
    if not result.scope_of_work: result.missing_information.append("Scope of work not identified confidently.")
    if not (result.inspection_factory_location or result.site_of_inspection): result.missing_information.append("Inspection location not identified confidently.")
    if not result.tests_or_checks: result.missing_information.append("No clear inspection activities / test items extracted.")
    if not result.applicable_standards: result.missing_information.append("No standards detected in the uploaded instruction. Use referenced ITP / QCP / project specifications for acceptance criteria.")
    return asdict(result)

def analyse_photo_placeholder(photo_name, inspection_area, inspection_item, document_context_raw):
    ctx = {}
    try: ctx = json.loads(document_context_raw or "{}")
    except Exception: ctx = {}
    standards = ctx.get("applicable_standards") or []
    standard_text = ", ".join(standards) if standards else "No explicit standard was detected in the instruction. Use the referenced ITP / QCP / project specification for exact acceptance criteria."
    return {"status_label": "Manual AI review required", "severity": "manual", "suggested_finding": "Vision model not connected yet", "inspection_area": inspection_area, "inspection_item": inspection_item, "reasoning": "The inspection workflow is ready to receive photos, but a real vision AI service is not connected yet. This response is intentionally not a fake defect decision.", "standard_reasoning": standard_text, "recommended_action": "Connect a vision model/API before using this result for pass/fail or defect classification."}

def analyse_file(path):
    p = Path(path)
    return analyse_text(extract_text(p), p.name)
